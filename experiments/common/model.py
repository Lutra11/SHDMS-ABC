"""Railway timetable model and dataset loaders shared by every experiment.

The loaders use workbook headers instead of fixed column positions wherever the
standardized dataset provides them.  This keeps preprocessing, experiments and
the manuscript tables tied to one data definition.
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, replace
from datetime import datetime, time
from pathlib import Path
from statistics import mean, median
from typing import Dict, List, Mapping, Sequence, Tuple

import numpy as np
from openpyxl import load_workbook

from .paths import DATASETS_DIR


@dataclass(frozen=True)
class DatasetSpec:
    code: str
    name: str
    directory: Path


@dataclass(frozen=True)
class PeriodDefinition:
    index: int
    label: str
    start: str
    end: str
    duration_min: int
    demand_share: float


@dataclass
class PeriodStats:
    count: int
    mean_interval: float
    median_interval: float
    min_interval: float
    max_interval: float
    baseline_departures: int


@dataclass
class StationCapacity:
    name: str
    km: float
    platforms: int
    lines: int
    base_dwell_min: float
    platform_buffer_min: float
    line_flow_coeff: float


@dataclass
class ModelParameters:
    train_capacity: int
    line_length_km: float
    run_cost_per_km: float = 18.0
    wait_cost_per_minute: float = 1.2
    average_ticket_fare: float = 12.0
    target_occupancy: float = 0.75
    load_min: float = 0.55
    load_max: float = 0.90
    safe_interval_min: float = 6.0
    board_dwell_coeff: float = 0.008
    alight_dwell_coeff: float = 0.006
    omega_run_cost: float = 0.35
    omega_wait_cost: float = 0.35
    omega_occ_loss: float = 0.15
    omega_stop_cost: float = 0.15
    penalty_cap: float = 2500.0
    penalty_load: float = 300.0
    penalty_revenue: float = 1200.0
    penalty_demand: float = 1800.0
    penalty_safe: float = 5000.0


@dataclass
class PlanMetrics:
    intervals: np.ndarray
    departures: np.ndarray
    dwell_times: np.ndarray
    reference_dwell_times: np.ndarray
    run_cost: float
    wait_cost: float
    occupancy_loss: float
    stop_adjustment_cost: float
    revenue: float
    total_demand: float
    served_demand: float
    demand_coverage: float
    weighted_interval: float
    average_waiting_time: float
    average_load: float
    max_capacity_utilization: float
    cap_violation: float
    load_violation: float
    revenue_violation: float
    demand_violation: float
    safe_violation: float
    objective: float
    platform_capacity: np.ndarray
    line_capacity: np.ndarray
    platform_utilization: np.ndarray
    line_utilization: np.ndarray
    combined_utilization: np.ndarray


@dataclass(frozen=True)
class ScenarioDefinition:
    code: str
    description: str
    overall_scale: float
    period_scale: Sequence[float]
    station_focus: Mapping[str, float]


DATASET_SPECS: Dict[str, DatasetSpec] = {
    "NST-HSR": DatasetSpec("NST-HSR", "NST-HSR", DATASETS_DIR / "NST-HSR"),
    "NCE-S": DatasetSpec("NCE-S", "New City Express-S", DATASETS_DIR / "Supplement" / "New City Express"),
    "WUX-S": DatasetSpec("WUX-S", "Wuxiao Intercity-S", DATASETS_DIR / "Supplement" / "Wuxiao Intercity"),
    "SSH-S": DatasetSpec("SSH-S", "Suishen Intercity-S", DATASETS_DIR / "Supplement" / "Suishen Intercity"),
    "BSC-S": DatasetSpec("BSC-S", "Beijing Sub-Center-S", DATASETS_DIR / "Supplement" / "Beijing Sub-Center"),
}


SCENARIOS: Dict[str, ScenarioDefinition] = {
    "S1": ScenarioDefinition("S1", "基准场景", 1.00, [1, 1, 1, 1, 1, 1], {}),
    "S2": ScenarioDefinition("S2", "整体需求增长场景", 1.10, [1, 1, 1, 1, 1, 1], {}),
    "S3": ScenarioDefinition("S3", "高峰强化场景", 1.12, [1, 1.20, 1, 1, 1.20, 1], {}),
    "S4": ScenarioDefinition(
        "S4", "关键车站压力测试场景", 1.20,
        [1, 1.15, 1, 1, 1.15, 1], {"南京南": 1.20, "江阴": 1.20, "太仓": 1.20},
    ),
}


def build_periods() -> List[PeriodDefinition]:
    return [
        PeriodDefinition(1, "06:00-08:00", "06:00", "08:00", 120, 0.09),
        PeriodDefinition(2, "08:00-10:00", "08:00", "10:00", 120, 0.20),
        PeriodDefinition(3, "10:00-14:00", "10:00", "14:00", 240, 0.19),
        PeriodDefinition(4, "14:00-17:00", "14:00", "17:00", 180, 0.16),
        PeriodDefinition(5, "17:00-20:00", "17:00", "20:00", 180, 0.24),
        PeriodDefinition(6, "20:00-22:00", "20:00", "22:00", 120, 0.12),
    ]


def normalize_station_name(value: object) -> str:
    text = re.sub(r"[（(].*?[)）]", "", str(value).strip())
    return text.replace("站", "").strip()


def to_minutes(value: object) -> int:
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)):
        fraction = float(value) % 1.0
        return int(round(fraction * 24 * 60))
    text = str(value).strip()
    for pattern in ("%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            pass
    raise ValueError(f"无法解析时间值: {value!r}")


def find_bundle_files(spec: DatasetSpec) -> Tuple[Path, Path, Path]:
    timetables = sorted(spec.directory.glob("1.*-Timetable.xlsx"))
    factors = sorted(spec.directory.glob("2.*-Factors.xlsx"))
    capacity = spec.directory / "3.Platform Capacity.xlsx"
    missing = []
    if not timetables:
        missing.append("1.*-Timetable.xlsx")
    if not factors:
        missing.append("2.*-Factors.xlsx")
    if not capacity.exists():
        missing.append("3.Platform Capacity.xlsx")
    if missing:
        raise FileNotFoundError(f"{spec.directory} 缺少: {', '.join(missing)}")
    return timetables[0], factors[0], capacity


def _sheet_rows(path: Path) -> Tuple[List[str], List[Tuple[object, ...]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(iterator)]
    return headers, list(iterator)


def load_timetable(path: Path) -> List[Dict[str, object]]:
    headers, values = _sheet_rows(path)
    col = {name: index for index, name in enumerate(headers)}
    required = ["Train code", "station1", "station2", "distance(km)", "time1", "time2", "interval(min)"]
    absent = [name for name in required if name not in col]
    if absent:
        raise ValueError(f"{path.name} 缺少字段: {', '.join(absent)}")

    def get(row: Tuple[object, ...], name: str, default: object = None) -> object:
        index = col.get(name)
        return row[index] if index is not None and index < len(row) else default

    rows: List[Dict[str, object]] = []
    for raw in values:
        if get(raw, "Train code") in (None, "") or get(raw, "interval(min)") is None:
            continue
        rows.append({
            "train_code": str(get(raw, "Train code")).strip(),
            "station1": normalize_station_name(get(raw, "station1")),
            "station2": normalize_station_name(get(raw, "station2")),
            "distance_km": float(get(raw, "distance(km)") or 0),
            "time1_min": to_minutes(get(raw, "time1")),
            "time2_min": to_minutes(get(raw, "time2")),
            "interval_min": float(get(raw, "interval(min)")),
            "second_class_fare": float(get(raw, "Second Class Fare", 0) or 0),
            "first_class_fare": float(get(raw, "First Class Fare", 0) or 0),
            "train_capacity": float(get(raw, "train_capacity", 0) or 0),
        })
    return rows


def load_station_capacities(path: Path) -> List[StationCapacity]:
    headers, rows = _sheet_rows(path)
    col = {name: index for index, name in enumerate(headers)}
    required = ["station_name", "km", "platforms", "lines"]
    absent = [name for name in required if name not in col]
    if absent:
        raise ValueError(f"{path.name} 缺少字段: {', '.join(absent)}")
    stations: List[StationCapacity] = []
    for row in rows:
        if row[col["station_name"]] in (None, ""):
            continue
        stations.append(StationCapacity(
            name=normalize_station_name(row[col["station_name"]]),
            km=float(row[col["km"]] or 0),
            platforms=int(row[col["platforms"]] or 2),
            lines=int(row[col["lines"]] or 4),
            base_dwell_min=float(row[col.get("base_dwell_min", -1)] or 2.2),
            platform_buffer_min=float(row[col.get("platform_buffer_min", -1)] or 10.0),
            line_flow_coeff=float(row[col.get("line_flow_coeff", -1)] or 0.020),
        ))
    if not stations:
        raise ValueError(f"{path.name} 没有有效车站记录")
    return stations


def load_factor_demands(
    path: Path, periods: Sequence[PeriodDefinition], stations: Sequence[StationCapacity]
) -> Tuple[np.ndarray, np.ndarray]:
    headers, rows = _sheet_rows(path)
    col = {name: index for index, name in enumerate(headers)}
    required = ["period_index", "station_name", "boarding_demand", "alighting_demand"]
    absent = [name for name in required if name not in col]
    if absent:
        raise ValueError(f"{path.name} 缺少字段: {', '.join(absent)}")
    period_index = {period.index: i for i, period in enumerate(periods)}
    station_index = {station.name: i for i, station in enumerate(stations)}
    board = np.zeros((len(periods), len(stations)), dtype=float)
    alight = np.zeros_like(board)
    seen = np.zeros_like(board, dtype=bool)
    for row in rows:
        if row[col["period_index"]] is None or row[col["station_name"]] is None:
            continue
        pi = period_index.get(int(row[col["period_index"]]))
        si = station_index.get(normalize_station_name(row[col["station_name"]]))
        if pi is None or si is None:
            continue
        board[pi, si] = float(row[col["boarding_demand"]] or 0)
        alight[pi, si] = float(row[col["alighting_demand"]] or 0)
        seen[pi, si] = True
    if not np.all(seen):
        missing = int(np.size(seen) - np.count_nonzero(seen))
        raise ValueError(f"{path.name} 缺少 {missing} 个时段-车站需求组合")
    return board, alight


def _match_period(periods: Sequence[PeriodDefinition], minutes: int) -> PeriodDefinition | None:
    for period in periods:
        if to_minutes(period.start) <= minutes < to_minutes(period.end):
            return period
    return None


def compute_period_statistics(
    rows: Sequence[Mapping[str, object]], periods: Sequence[PeriodDefinition]
) -> Dict[int, PeriodStats]:
    intervals: Dict[int, List[float]] = {period.index: [] for period in periods}
    for row in rows:
        period = _match_period(periods, int(row["time1_min"]))
        if period is not None:
            intervals[period.index].append(float(row["interval_min"]))
    result: Dict[int, PeriodStats] = {}
    for period in periods:
        values = intervals[period.index]
        if not values:
            raise ValueError(f"时段 {period.label} 没有有效发车间隔记录")
        avg = float(mean(values))
        result[period.index] = PeriodStats(
            len(values), avg, float(median(values)), min(values), max(values),
            max(1, round(period.duration_min / avg)),
        )
    return result


class HSRModel:
    def __init__(self, periods, stations, period_stats, board_demand, alight_demand, parameters):
        self.periods = list(periods)
        self.stations = list(stations)
        self.period_stats = period_stats
        self.board_demand = np.asarray(board_demand, dtype=float)
        self.alight_demand = np.asarray(alight_demand, dtype=float)
        self.parameters = parameters
        self.period_count = len(periods)
        self.station_count = len(stations)
        self.base_dwell = np.array([s.base_dwell_min for s in stations], dtype=float)
        self.platform_buffer = np.array([s.platform_buffer_min for s in stations], dtype=float)
        self.platforms = np.array([s.platforms for s in stations], dtype=float)
        self.lines = np.array([s.lines for s in stations], dtype=float)
        self.line_flow = np.array([s.line_flow_coeff for s in stations], dtype=float)
        self.durations = np.array([p.duration_min for p in periods], dtype=float)
        self.baseline_intervals = np.array([period_stats[p.index].mean_interval for p in periods])
        self.interval_lower = np.maximum(parameters.safe_interval_min, self.baseline_intervals * 0.70)
        self.interval_upper = np.minimum(
            np.maximum(self.baseline_intervals * 1.35, self.interval_lower + 2), 30.0
        )
        self.dwell_lower = np.full((self.period_count, self.station_count), 1.5)
        self.dwell_upper = np.full((self.period_count, self.station_count), 8.0)
        departure = np.maximum(1, np.rint(self.durations / self.baseline_intervals)).astype(int)
        dwell = self.reference_dwell_times(self.baseline_intervals, departure)
        baseline = self._evaluate_raw(self.baseline_intervals, dwell)
        self._scales = {
            "run": max(baseline.run_cost, 1.0), "wait": max(baseline.wait_cost, 1.0),
            "occ": max(baseline.occupancy_loss, 1.0), "stop": max(baseline.stop_adjustment_cost, 1.0),
        }

    def bounds(self) -> List[Tuple[float, float]]:
        return list(zip(self.interval_lower, self.interval_upper)) + list(zip(self.dwell_lower.ravel(), self.dwell_upper.ravel()))

    def vector_to_plan(self, vector: Sequence[float]) -> Tuple[np.ndarray, np.ndarray]:
        values = np.asarray(vector, dtype=float)
        intervals = np.clip(values[:self.period_count], self.interval_lower, self.interval_upper)
        dwell = values[self.period_count:].reshape(self.period_count, self.station_count)
        return intervals, np.clip(dwell, self.dwell_lower, self.dwell_upper)

    def reference_dwell_times(self, intervals: np.ndarray, departures: np.ndarray) -> np.ndarray:
        reference = self.base_dwell[None, :] + self.parameters.board_dwell_coeff * self.board_demand / departures[:, None] + self.parameters.alight_dwell_coeff * self.alight_demand / departures[:, None]
        return np.clip(reference, self.dwell_lower, self.dwell_upper)

    def baseline_vector(self) -> np.ndarray:
        departures = np.maximum(1, np.rint(self.durations / self.baseline_intervals)).astype(int)
        dwell = self.reference_dwell_times(self.baseline_intervals, departures)
        return np.concatenate([self.baseline_intervals, dwell.ravel()])

    def heuristic_vector(self) -> np.ndarray:
        demand = self.board_demand.sum(axis=1)
        intervals = np.clip(self.baseline_intervals * (demand / max(float(np.mean(demand)), 1)) ** -0.28, self.interval_lower, self.interval_upper)
        departures = np.maximum(1, np.rint(self.durations / intervals)).astype(int)
        dwell = self.reference_dwell_times(intervals, departures)
        return np.concatenate([intervals, dwell.ravel()])

    def objective(self, vector: Sequence[float]) -> float:
        intervals, dwell = self.vector_to_plan(vector)
        return self._evaluate_raw(intervals, dwell).objective

    def evaluate_vector(self, vector: Sequence[float]) -> PlanMetrics:
        intervals, dwell = self.vector_to_plan(vector)
        return self._evaluate_raw(intervals, dwell)

    def _poisson_target_loss(self, expected: float) -> float:
        target = self.parameters.target_occupancy * self.parameters.train_capacity
        if expected <= 1e-9:
            return 1.0
        if expected > 30:
            z = (target - 0.5 - expected) / math.sqrt(expected)
            probability = 1 - 0.5 * (1 + math.erf(z / math.sqrt(2)))
            return max(0.0, min(1.0, 1 - probability))
        mass = math.exp(-expected)
        cumulative = mass
        for passengers in range(1, int(target)):
            mass *= expected / passengers
            cumulative += mass
        return max(0.0, min(1.0, cumulative))

    def _evaluate_raw(self, intervals: np.ndarray, dwell_times: np.ndarray) -> PlanMetrics:
        p = self.parameters
        departures = np.maximum(1, np.rint(self.durations / intervals)).astype(int)
        reference = self.reference_dwell_times(intervals, departures)
        dwell = np.clip(dwell_times, self.dwell_lower, self.dwell_upper)
        demand = self.board_demand.sum(axis=1)
        train_capacity = p.train_capacity * departures
        served = np.minimum(demand, train_capacity)
        wait_cost = float(0.5 * p.wait_cost_per_minute * np.sum(self.board_demand * intervals[:, None]))
        run_cost = float(p.run_cost_per_km * p.line_length_km * np.sum(departures))
        occupancy_loss = float(sum(self._poisson_target_loss(demand[i] / departures[i]) for i in range(self.period_count)))
        stop_cost = float(np.sum((dwell - reference) ** 2))
        platform_capacity = self.platforms[None, :] * self.durations[:, None] / (dwell + self.platform_buffer[None, :])
        line_capacity = self.line_flow[None, :] * self.lines[None, :] * self.durations[:, None]
        combined_capacity = np.minimum(platform_capacity, line_capacity)
        platform_util = departures[:, None] / np.maximum(platform_capacity, 1e-9)
        line_util = departures[:, None] / np.maximum(line_capacity, 1e-9)
        combined_util = departures[:, None] / np.maximum(combined_capacity, 1e-9)
        cap_violation = float(np.sum(np.maximum(0, combined_util - 1) ** 2))
        load = demand / np.maximum(train_capacity, 1)
        load_violation = float(np.sum(np.maximum(0, p.load_min - load) ** 2) + np.sum(np.maximum(0, load - p.load_max) ** 2))
        revenue = float(p.average_ticket_fare * np.sum(served))
        revenue_violation = float(max(0, (run_cost - revenue) / max(run_cost, 1)) ** 2)
        demand_violation = float(np.sum((np.maximum(0, demand - train_capacity) / np.maximum(demand, 1)) ** 2))
        safe_violation = float(np.sum((np.maximum(0, p.safe_interval_min - intervals) / p.safe_interval_min) ** 2))
        total_demand, served_demand = float(np.sum(demand)), float(np.sum(served))
        weighted_interval = float(np.sum(demand * intervals) / max(total_demand, 1))
        scales = getattr(self, "_scales", {"run": max(run_cost, 1), "wait": max(wait_cost, 1), "occ": max(occupancy_loss, 1), "stop": max(stop_cost, 1)})
        objective = (
            p.omega_run_cost * run_cost / scales["run"] + p.omega_wait_cost * wait_cost / scales["wait"]
            + p.omega_occ_loss * occupancy_loss / scales["occ"] + p.omega_stop_cost * stop_cost / scales["stop"]
            + p.penalty_cap * cap_violation + p.penalty_load * load_violation
            + p.penalty_revenue * revenue_violation + p.penalty_demand * demand_violation + p.penalty_safe * safe_violation
        )
        return PlanMetrics(
            intervals.copy(), departures.copy(), dwell.copy(), reference.copy(), run_cost, wait_cost,
            occupancy_loss, stop_cost, revenue, total_demand, served_demand, served_demand / max(total_demand, 1),
            weighted_interval, weighted_interval / 2, served_demand / max(float(np.sum(train_capacity)), 1),
            float(np.max(combined_util)), cap_violation, load_violation, revenue_violation, demand_violation,
            safe_violation, float(objective), platform_capacity, line_capacity, platform_util, line_util, combined_util,
        )


class ImprovedModel:
    def __init__(self, base_model: HSRModel):
        self.base = base_model

    def bounds(self):
        return self.base.bounds()

    def baseline_vector(self):
        return self.base.baseline_vector()

    def heuristic_vector(self):
        return self.base.heuristic_vector()

    def seed_solutions(self) -> List[List[float]]:
        baseline, heuristic = self.baseline_vector(), self.heuristic_vector()
        midpoint = 0.35 * baseline + 0.65 * heuristic
        aggressive = heuristic.copy()
        aggressive[:self.base.period_count] = np.clip(aggressive[:self.base.period_count] * 0.97, self.base.interval_lower, self.base.interval_upper)
        return [baseline.tolist(), heuristic.tolist(), midpoint.tolist(), aggressive.tolist()]

    def repair_vector(self, vector: Sequence[float]) -> np.ndarray:
        intervals, dwell = self.base.vector_to_plan(vector)
        intervals, dwell = intervals.copy(), dwell.copy()
        for _ in range(8):
            departures = np.maximum(1, np.rint(self.base.durations / intervals)).astype(int)
            dwell = np.clip(np.minimum(dwell, self.base.reference_dwell_times(intervals, departures)), self.base.dwell_lower, self.base.dwell_upper)
            platform = self.base.platforms[None, :] * self.base.durations[:, None] / (dwell + self.base.platform_buffer[None, :])
            line = self.base.line_flow[None, :] * self.base.lines[None, :] * self.base.durations[:, None]
            allowable = np.clip(np.floor(np.min(np.minimum(platform, line), axis=1) + 1e-9).astype(int), 1, None)
            new_departures = np.minimum(departures, allowable)
            if np.array_equal(new_departures, departures):
                break
            intervals = np.clip(self.base.durations / new_departures, self.base.interval_lower, self.base.interval_upper)
        return np.concatenate([intervals, dwell.ravel()])

    def objective(self, vector: Sequence[float]) -> float:
        return float(self.base.objective(vector))

    def evaluate_vector(self, vector: Sequence[float]) -> PlanMetrics:
        return self.base.evaluate_vector(self.repair_vector(vector))


def build_dataset_model(code: str = "NST-HSR", parameter_overrides: Mapping[str, float] | None = None) -> ImprovedModel:
    if code not in DATASET_SPECS:
        raise KeyError(f"未知数据集 {code!r}；可选值: {', '.join(DATASET_SPECS)}")
    timetable, factors, capacity = find_bundle_files(DATASET_SPECS[code])
    rows = load_timetable(timetable)
    periods = build_periods()
    stations = load_station_capacities(capacity)
    board, alight = load_factor_demands(factors, periods, stations)
    capacities = [float(row["train_capacity"]) for row in rows if float(row["train_capacity"]) > 0]
    fares = [float(row["second_class_fare"]) for row in rows if float(row["second_class_fare"]) > 0]
    parameters = ModelParameters(
        train_capacity=int(round(median(capacities))) if capacities else 600,
        line_length_km=max(station.km for station in stations),
        average_ticket_fare=float(mean(fares)) if fares else 12.0,
    )
    if parameter_overrides:
        parameters = replace(parameters, **dict(parameter_overrides))
    return ImprovedModel(HSRModel(periods, stations, compute_period_statistics(rows, periods), board, alight, parameters))


def create_scenario_model(model: ImprovedModel, scenario: str | ScenarioDefinition) -> ImprovedModel:
    definition = SCENARIOS[scenario] if isinstance(scenario, str) else scenario
    base = model.base
    station_scale = np.array([definition.station_focus.get(s.name, 1.0) for s in base.stations])
    period_scale = np.asarray(definition.period_scale, dtype=float)[:, None]
    board = base.board_demand * period_scale * station_scale[None, :]
    alight = base.alight_demand * period_scale * station_scale[None, :]
    target = float(np.sum(base.board_demand) * definition.overall_scale)
    board *= target / max(float(np.sum(board)), 1e-12)
    alight *= target / max(float(np.sum(alight)), 1e-12)
    return ImprovedModel(HSRModel(base.periods, base.stations, base.period_stats, board, alight, base.parameters))


def hard_violation(metrics: PlanMetrics) -> float:
    return max(0.0, metrics.max_capacity_utilization - 1.0) + metrics.safe_violation


def soft_violation(metrics: PlanMetrics) -> float:
    return metrics.demand_violation + 0.5 * metrics.load_violation + metrics.revenue_violation


def benchmark_score(metrics: PlanMetrics) -> float:
    return metrics.objective + 1.0e5 * hard_violation(metrics) + 1.0e3 * soft_violation(metrics)


def benchmark_rank_key(metrics: PlanMetrics) -> tuple:
    return (hard_violation(metrics) > 1e-9, hard_violation(metrics), soft_violation(metrics), metrics.objective, metrics.average_waiting_time)
