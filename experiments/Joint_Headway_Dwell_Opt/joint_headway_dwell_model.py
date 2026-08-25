"""Canonical joint optimization model for train headways and dwell times.

This module is the executable counterpart of manuscript Sections 3.1.1--3.1.4.
It provides:

* the four-stage passenger-demand equations (1)--(10);
* a configurable non-homogeneous Poisson arrival model (11)--(13);
* headway and station-dwell decision encoding;
* operating, waiting, occupancy and dwell-adjustment costs (14)--(18);
* station-capacity, load, revenue, demand and safety constraints (19)--(26);
* the composite objective in equation (27); and
* dataset/scenario builders used by every Chapter 4 experiment.

The current workbooks already contain the final period-by-station demand from
equation (10).  ``build_dataset_model`` therefore loads those values directly,
while ``PassengerDemandModel`` exposes equations (1)--(10) for datasets that
provide the required historical OD and modal attributes.
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, replace
from datetime import datetime, time
from pathlib import Path
from statistics import mean, median
from typing import Callable, Dict, List, Mapping, Sequence, Tuple

import numpy as np
from openpyxl import load_workbook


EXPERIMENTS_DIR = Path(__file__).resolve().parent.parent
GIT_CONTENT_DIR = EXPERIMENTS_DIR.parent
DATASETS_DIR = GIT_CONTENT_DIR / "datasets"


@dataclass(frozen=True)
class DatasetSpec:
    code: str
    name: str
    directory: Path


@dataclass(frozen=True)
class PeriodDefinition:
    index: int
    label: str
    start: str
    end: str
    duration_min: int
    demand_share: float


@dataclass
class PeriodStats:
    count: int
    mean_interval: float
    median_interval: float
    min_interval: float
    max_interval: float
    baseline_departures: int


@dataclass
class StationCapacity:
    name: str
    km: float
    platforms: int
    lines: int
    base_dwell_min: float
    platform_buffer_min: float
    line_flow_coeff: float
    min_dwell_min: float = 1.5
    max_dwell_min: float = 8.0
    station_class: str = "standard"


@dataclass(frozen=True)
class GeneralizedCostCoefficients:
    """Coefficients B1--B5 in manuscript equation (1)."""

    intercept: float
    direct_cost: float
    travel_time: float
    station_access_time: float
    waiting_time: float


@dataclass
class ModelParameters:
    train_capacity: int
    line_length_km: float
    run_cost_per_km: float = 18.0
    wait_cost_per_minute: float = 1.2
    average_ticket_fare: float = 12.0
    target_occupancy: float = 0.75
    load_min: float = 0.55
    load_max: float = 0.90
    safe_interval_min: float = 6.0
    board_dwell_coeff: float = 0.008
    alight_dwell_coeff: float = 0.006
    omega_run_cost: float = 0.35
    omega_wait_cost: float = 0.35
    omega_occ_loss: float = 0.15
    omega_stop_cost: float = 0.15
    penalty_cap: float = 2500.0
    penalty_load: float = 300.0
    penalty_revenue: float = 1200.0
    penalty_demand: float = 1800.0
    penalty_safe: float = 5000.0
    epsilon: float = 1.0e-9
    arrival_concentration: float = 0.0
    normalize_objectives: bool = True
    normalize_violations: bool = True


@dataclass
class PlanMetrics:
    intervals: np.ndarray
    departures: np.ndarray
    dwell_times: np.ndarray
    reference_dwell_times: np.ndarray
    run_cost: float
    wait_cost: float
    occupancy_loss: float
    stop_adjustment_cost: float
    revenue: float
    total_demand: float
    served_demand: float
    demand_coverage: float
    weighted_interval: float
    average_waiting_time: float
    average_load: float
    max_capacity_utilization: float
    cap_violation: float
    load_violation: float
    revenue_violation: float
    demand_violation: float
    safe_violation: float
    objective: float
    platform_capacity: np.ndarray
    line_capacity: np.ndarray
    platform_utilization: np.ndarray
    line_utilization: np.ndarray
    combined_utilization: np.ndarray


@dataclass(frozen=True)
class ScenarioDefinition:
    code: str
    description: str
    overall_scale: float
    period_scale: Sequence[float]
    station_focus: Mapping[int, float]


DATASET_SPECS: Dict[str, DatasetSpec] = {
    "NST-HSR": DatasetSpec("NST-HSR", "NST-HSR", DATASETS_DIR / "NST-HSR"),
    "NCE-S": DatasetSpec("NCE-S", "New City Express-S", DATASETS_DIR / "Supplement" / "New City Express"),
    "WUX-S": DatasetSpec("WUX-S", "Wuxiao Intercity-S", DATASETS_DIR / "Supplement" / "Wuxiao Intercity"),
    "SSH-S": DatasetSpec("SSH-S", "Suishen Intercity-S", DATASETS_DIR / "Supplement" / "Suishen Intercity"),
    "BSC-S": DatasetSpec("BSC-S", "Beijing Sub-Center-S", DATASETS_DIR / "Supplement" / "Beijing Sub-Center"),
}


SCENARIOS: Dict[str, ScenarioDefinition] = {
    "S1": ScenarioDefinition("S1", "Baseline scenario", 1.00, [1, 1, 1, 1, 1, 1], {}),
    "S2": ScenarioDefinition("S2", "Overall demand growth", 1.10, [1, 1, 1, 1, 1, 1], {}),
    "S3": ScenarioDefinition("S3", "Peak demand intensification", 1.12, [1, 1.20, 1, 1, 1.20, 1], {}),
    "S4": ScenarioDefinition(
        "S4", "Key-station stress test", 1.20,
        [1, 1.15, 1, 1, 1.15, 1], {0: 1.20, 5: 1.20, 8: 1.20},
    ),
}


def _as_finite_array(value: object, name: str) -> np.ndarray:
    array = np.asarray(value, dtype=float)
    if not np.all(np.isfinite(array)):
        raise ValueError(f"{name} must contain only finite values")
    return array


def generalized_cost(
    direct_cost: object,
    travel_time: object,
    station_access_time: object,
    waiting_time: object,
    coefficients: GeneralizedCostCoefficients,
) -> np.ndarray:
    """Return generalized travel cost U_ij^m from equation (1)."""

    cost = _as_finite_array(direct_cost, "direct_cost")
    travel = _as_finite_array(travel_time, "travel_time")
    access = _as_finite_array(station_access_time, "station_access_time")
    waiting = _as_finite_array(waiting_time, "waiting_time")
    return (
        coefficients.intercept
        + coefficients.direct_cost * cost
        + coefficients.travel_time * travel
        + coefficients.station_access_time * access
        + coefficients.waiting_time * waiting
    )


def gm11_forecast(history: Sequence[float]) -> float:
    """Forecast the next direct-demand observation using GM(1,1), equations (2)--(4)."""

    x0 = _as_finite_array(history, "history").reshape(-1)
    if x0.size < 4:
        raise ValueError("GM(1,1) requires at least four observations")
    if np.any(x0 <= 0):
        raise ValueError("GM(1,1) observations must be positive")
    x1 = np.cumsum(x0)
    z1 = 0.5 * (x1[1:] + x1[:-1])
    matrix = np.column_stack((-z1, np.ones_like(z1)))
    a, b = np.linalg.lstsq(matrix, x0[1:], rcond=None)[0]
    if abs(a) <= 1.0e-12:
        forecast = float(x0[-1] + b)
    else:
        constant = x0[0] - b / a
        accumulated_n = constant * math.exp(-a * (x0.size - 1)) + b / a
        accumulated_next = constant * math.exp(-a * x0.size) + b / a
        forecast = accumulated_next - accumulated_n
    return max(0.0, float(forecast))


def induced_demand(base_demand: object, cost_before: object, cost_after: object, gamma: float) -> np.ndarray:
    """Apply the induced-rate model in equations (5)--(6)."""

    base = _as_finite_array(base_demand, "base_demand")
    before = _as_finite_array(cost_before, "cost_before")
    after = _as_finite_array(cost_after, "cost_after")
    if np.any(base < 0) or np.any(before <= 0) or np.any(after <= 0):
        raise ValueError("demand must be non-negative and generalized costs must be positive")
    induced_rate = np.power(after / before, float(gamma)) - 1.0
    return np.maximum(0.0, base * (1.0 + induced_rate))


def logit_mode_probabilities(generalized_costs: object, axis: int = -1) -> np.ndarray:
    """Return numerically stable Logit mode shares from equation (7)."""

    costs = _as_finite_array(generalized_costs, "generalized_costs")
    shifted = -costs - np.max(-costs, axis=axis, keepdims=True)
    exp_values = np.exp(shifted)
    return exp_values / np.sum(exp_values, axis=axis, keepdims=True)


def allocate_rail_demand(total_demand: object, rail_probability: object) -> np.ndarray:
    """Allocate total OD demand to rail using equation (8)."""

    total = _as_finite_array(total_demand, "total_demand")
    probability = _as_finite_array(rail_probability, "rail_probability")
    if np.any(total < 0) or np.any((probability < 0) | (probability > 1)):
        raise ValueError("demand must be non-negative and probabilities must lie in [0, 1]")
    return total * probability


def map_od_demand_to_period_station(
    rail_od_demand: object,
    period_od_shares: object,
) -> Tuple[np.ndarray, np.ndarray]:
    """Map OD rail demand to period/station boarding and alighting, equations (9)--(10)."""

    od = _as_finite_array(rail_od_demand, "rail_od_demand")
    shares = _as_finite_array(period_od_shares, "period_od_shares")
    if od.ndim != 2 or od.shape[0] != od.shape[1]:
        raise ValueError("rail_od_demand must be a square station-by-station matrix")
    if shares.ndim != 3 or shares.shape[1:] != od.shape:
        raise ValueError("period_od_shares must have shape (periods, stations, stations)")
    if np.any(od < 0) or np.any(shares < 0):
        raise ValueError("OD demand and period shares must be non-negative")
    active = od > 0
    if np.any(active) and not np.allclose(np.sum(shares, axis=0)[active], 1.0, atol=1.0e-8):
        raise ValueError("period shares must sum to one for every active OD pair")
    allocated = shares * od[None, :, :]
    period_count, station_count, _ = allocated.shape
    board = np.zeros((period_count, station_count), dtype=float)
    alight = np.zeros_like(board)
    for station in range(station_count):
        board[:, station] = np.sum(allocated[:, station, station + 1 :], axis=1)
        alight[:, station] = np.sum(allocated[:, :station, station], axis=1)
    return board, alight


class PassengerDemandModel:
    """Named access to the complete demand pipeline in manuscript Section 3.1.2."""

    generalized_cost = staticmethod(generalized_cost)
    direct_demand_forecast = staticmethod(gm11_forecast)
    induced_demand = staticmethod(induced_demand)
    mode_probabilities = staticmethod(logit_mode_probabilities)
    rail_demand = staticmethod(allocate_rail_demand)
    period_station_demand = staticmethod(map_od_demand_to_period_station)


def power_arrival_intensity(
    time_from_previous_departure: object,
    headway: float,
    expected_arrivals: float,
    concentration: float = 0.0,
) -> np.ndarray:
    """NHPP intensity lambda(t) for equation (11).

    ``concentration=0`` is the uniform-arrival special case. Positive values
    move more arrivals towards the scheduled departure while preserving the
    expected number of arrivals over the headway.
    """

    if headway <= 0 or expected_arrivals < 0 or concentration < 0:
        raise ValueError("headway must be positive; arrivals and concentration must be non-negative")
    values = _as_finite_array(time_from_previous_departure, "time_from_previous_departure")
    if np.any((values < 0) | (values > headway)):
        raise ValueError("arrival times must lie inside the headway")
    scaled = values / headway
    return expected_arrivals * (concentration + 1.0) * np.power(scaled, concentration) / headway


def cumulative_arrival_intensity(
    start: float,
    end: float,
    headway: float,
    expected_arrivals: float,
    concentration: float = 0.0,
) -> float:
    """Return Lambda(a,b), the cumulative intensity in equation (11)."""

    if not 0 <= start <= end <= headway:
        raise ValueError("the integration interval must satisfy 0 <= start <= end <= headway")
    exponent = concentration + 1.0
    return float(expected_arrivals * ((end / headway) ** exponent - (start / headway) ** exponent))


def integrate_interval_waiting_time(
    arrival_intensity: Callable[[np.ndarray], np.ndarray],
    headway: float,
    grid_points: int = 257,
) -> float:
    """Numerically evaluate equation (12) for an arbitrary lambda(t)."""

    if headway <= 0 or grid_points < 3:
        raise ValueError("headway must be positive and grid_points must be at least three")
    grid = np.linspace(0.0, headway, int(grid_points))
    intensity = _as_finite_array(arrival_intensity(grid), "arrival_intensity")
    if intensity.shape != grid.shape or np.any(intensity < 0):
        raise ValueError("arrival_intensity must return one non-negative value per grid point")
    return float(np.trapezoid((headway - grid) * intensity, grid))


def nhpp_mean_waiting_time(headway: object, concentration: float = 0.0) -> np.ndarray:
    """Closed-form expected wait for the power NHPP used in equations (12)--(13)."""

    intervals = _as_finite_array(headway, "headway")
    if np.any(intervals <= 0) or concentration < 0:
        raise ValueError("headways must be positive and concentration must be non-negative")
    return intervals / (float(concentration) + 2.0)


def build_periods() -> List[PeriodDefinition]:
    return [
        PeriodDefinition(1, "06:00-08:00", "06:00", "08:00", 120, 0.09),
        PeriodDefinition(2, "08:00-10:00", "08:00", "10:00", 120, 0.20),
        PeriodDefinition(3, "10:00-14:00", "10:00", "14:00", 240, 0.19),
        PeriodDefinition(4, "14:00-17:00", "14:00", "17:00", 180, 0.16),
        PeriodDefinition(5, "17:00-20:00", "17:00", "20:00", 180, 0.24),
        PeriodDefinition(6, "20:00-22:00", "20:00", "22:00", 120, 0.12),
    ]


def normalize_station_name(value: object) -> str:
    text = re.sub(r"[\uFF08(].*?[)\uFF09]", "", str(value).strip())
    station_suffix = chr(0x7AD9)
    return text.removesuffix(station_suffix).strip()


def to_minutes(value: object) -> int:
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)):
        fraction = float(value) % 1.0
        return int(round(fraction * 24 * 60))
    text = str(value).strip()
    for pattern in ("%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            pass
    raise ValueError(f"Cannot parse time value: {value!r}")


def find_bundle_files(spec: DatasetSpec) -> Tuple[Path, Path, Path]:
    timetables = sorted(spec.directory.glob("1.*-Timetable.xlsx"))
    factors = sorted(spec.directory.glob("2.*-Factors.xlsx"))
    capacity = spec.directory / "3.Platform Capacity.xlsx"
    missing = []
    if not timetables:
        missing.append("1.*-Timetable.xlsx")
    if not factors:
        missing.append("2.*-Factors.xlsx")
    if not capacity.exists():
        missing.append("3.Platform Capacity.xlsx")
    if missing:
        raise FileNotFoundError(f"{spec.directory} is missing: {', '.join(missing)}")
    return timetables[0], factors[0], capacity


def _sheet_rows(path: Path) -> Tuple[List[str], List[Tuple[object, ...]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
        return headers, list(iterator)
    finally:
        workbook.close()


def load_timetable(path: Path) -> List[Dict[str, object]]:
    headers, values = _sheet_rows(path)
    col = {name: index for index, name in enumerate(headers)}
    required = ["Train code", "station1", "station2", "distance(km)", "time1", "time2", "interval(min)"]
    absent = [name for name in required if name not in col]
    if absent:
        raise ValueError(f"{path.name} is missing fields: {', '.join(absent)}")

    def get(row: Tuple[object, ...], name: str, default: object = None) -> object:
        index = col.get(name)
        return row[index] if index is not None and index < len(row) else default

    rows: List[Dict[str, object]] = []
    for raw in values:
        if get(raw, "Train code") in (None, "") or get(raw, "interval(min)") is None:
            continue
        rows.append({
            "train_code": str(get(raw, "Train code")).strip(),
            "station1": normalize_station_name(get(raw, "station1")),
            "station2": normalize_station_name(get(raw, "station2")),
            "distance_km": float(get(raw, "distance(km)") or 0),
            "time1_min": to_minutes(get(raw, "time1")),
            "time2_min": to_minutes(get(raw, "time2")),
            "interval_min": float(get(raw, "interval(min)")),
            "second_class_fare": float(get(raw, "Second Class Fare", 0) or 0),
            "first_class_fare": float(get(raw, "First Class Fare", 0) or 0),
            "train_capacity": float(get(raw, "train_capacity", 0) or 0),
        })
    return rows


def load_station_capacities(path: Path) -> List[StationCapacity]:
    headers, rows = _sheet_rows(path)
    col = {name: index for index, name in enumerate(headers)}
    required = ["station_name", "km", "platforms", "lines"]
    absent = [name for name in required if name not in col]
    if absent:
        raise ValueError(f"{path.name} is missing fields: {', '.join(absent)}")

    def get(row: Tuple[object, ...], name: str, default: object) -> object:
        index = col.get(name)
        if index is None or index >= len(row) or row[index] in (None, ""):
            return default
        return row[index]

    stations: List[StationCapacity] = []
    for row in rows:
        if get(row, "station_name", None) in (None, ""):
            continue
        stations.append(StationCapacity(
            name=normalize_station_name(get(row, "station_name", "")),
            km=float(get(row, "km", 0.0)),
            platforms=int(get(row, "platforms", 2)),
            lines=int(get(row, "lines", 4)),
            base_dwell_min=float(get(row, "base_dwell_min", 2.2)),
            platform_buffer_min=float(get(row, "platform_buffer_min", 10.0)),
            line_flow_coeff=float(get(row, "line_flow_coeff", 0.020)),
            min_dwell_min=float(get(row, "min_dwell_min", 1.5)),
            max_dwell_min=float(get(row, "max_dwell_min", 8.0)),
            station_class=str(get(row, "station_class", "standard")),
        ))
    if not stations:
        raise ValueError(f"{path.name} contains no valid station records")
    for station in stations:
        if not 0 <= station.min_dwell_min <= station.base_dwell_min <= station.max_dwell_min:
            raise ValueError(f"Invalid dwell-time bounds for {station.name}")
    return stations


def load_factor_demands(
    path: Path,
    periods: Sequence[PeriodDefinition],
    stations: Sequence[StationCapacity],
) -> Tuple[np.ndarray, np.ndarray]:
    headers, rows = _sheet_rows(path)
    col = {name: index for index, name in enumerate(headers)}
    required = ["period_index", "station_name", "boarding_demand", "alighting_demand"]
    absent = [name for name in required if name not in col]
    if absent:
        raise ValueError(f"{path.name} is missing fields: {', '.join(absent)}")
    period_index = {period.index: index for index, period in enumerate(periods)}
    station_index = {station.name: index for index, station in enumerate(stations)}
    board = np.zeros((len(periods), len(stations)), dtype=float)
    alight = np.zeros_like(board)
    seen = np.zeros_like(board, dtype=bool)
    for row in rows:
        if row[col["period_index"]] is None or row[col["station_name"]] is None:
            continue
        pi = period_index.get(int(row[col["period_index"]]))
        si = station_index.get(normalize_station_name(row[col["station_name"]]))
        if pi is None or si is None:
            continue
        board[pi, si] = float(row[col["boarding_demand"]] or 0)
        alight[pi, si] = float(row[col["alighting_demand"]] or 0)
        seen[pi, si] = True
    if not np.all(seen):
        missing = int(np.size(seen) - np.count_nonzero(seen))
        raise ValueError(f"{path.name} is missing {missing} period-station demand combinations")
    if np.any(board < 0) or np.any(alight < 0):
        raise ValueError(f"{path.name} contains negative passenger demand")
    return board, alight


def _match_period(periods: Sequence[PeriodDefinition], minutes: int) -> PeriodDefinition | None:
    for period in periods:
        if to_minutes(period.start) <= minutes < to_minutes(period.end):
            return period
    return None


def compute_period_statistics(
    rows: Sequence[Mapping[str, object]], periods: Sequence[PeriodDefinition]
) -> Dict[int, PeriodStats]:
    intervals: Dict[int, List[float]] = {period.index: [] for period in periods}
    for row in rows:
        period = _match_period(periods, int(row["time1_min"]))
        if period is not None:
            intervals[period.index].append(float(row["interval_min"]))
    result: Dict[int, PeriodStats] = {}
    for period in periods:
        values = intervals[period.index]
        if not values:
            raise ValueError(f"Period {period.label} has no valid headway records")
        average = float(mean(values))
        result[period.index] = PeriodStats(
            len(values), average, float(median(values)), min(values), max(values),
            max(1, round(period.duration_min / average)),
        )
    return result


class JointHeadwayDwellModel:
    """Executable constrained model corresponding to Sections 3.1.1--3.1.4."""

    def __init__(
        self,
        periods: Sequence[PeriodDefinition],
        stations: Sequence[StationCapacity],
        period_stats: Mapping[int, PeriodStats],
        board_demand: object,
        alight_demand: object,
        parameters: ModelParameters,
        stop_pattern: object | None = None,
    ):
        self.periods = list(periods)
        self.stations = list(stations)
        self.period_stats = dict(period_stats)
        self.parameters = parameters
        self.period_count = len(self.periods)
        self.station_count = len(self.stations)
        expected_shape = (self.period_count, self.station_count)
        self.board_demand = _as_finite_array(board_demand, "board_demand")
        self.alight_demand = _as_finite_array(alight_demand, "alight_demand")
        if self.board_demand.shape != expected_shape or self.alight_demand.shape != expected_shape:
            raise ValueError(f"boarding and alighting demand must have shape {expected_shape}")
        if np.any(self.board_demand < 0) or np.any(self.alight_demand < 0):
            raise ValueError("passenger demand must be non-negative")
        if stop_pattern is None:
            self.stop_pattern = np.ones(expected_shape, dtype=float)
        else:
            self.stop_pattern = _as_finite_array(stop_pattern, "stop_pattern")
            if self.stop_pattern.shape != expected_shape or np.any(~np.isin(self.stop_pattern, [0.0, 1.0])):
                raise ValueError(f"stop_pattern must be a binary matrix with shape {expected_shape}")
        self.base_dwell = np.array([station.base_dwell_min for station in self.stations], dtype=float)
        self.platform_buffer = np.array([station.platform_buffer_min for station in self.stations], dtype=float)
        self.platforms = np.array([station.platforms for station in self.stations], dtype=float)
        self.lines = np.array([station.lines for station in self.stations], dtype=float)
        self.line_flow = np.array([station.line_flow_coeff for station in self.stations], dtype=float)
        self.durations = np.array([period.duration_min for period in self.periods], dtype=float)
        self.baseline_intervals = np.array(
            [self.period_stats[period.index].mean_interval for period in self.periods], dtype=float
        )
        self.interval_lower = np.maximum(parameters.safe_interval_min, self.baseline_intervals * 0.70)
        self.interval_upper = np.minimum(
            np.maximum(self.baseline_intervals * 1.35, self.interval_lower + 2.0), 30.0
        )
        station_lower = np.array([station.min_dwell_min for station in self.stations], dtype=float)
        station_upper = np.array([station.max_dwell_min for station in self.stations], dtype=float)
        self.dwell_lower = self.stop_pattern * station_lower[None, :]
        self.dwell_upper = self.stop_pattern * station_upper[None, :]
        if not np.all(self.dwell_lower <= self.dwell_upper):
            raise ValueError("station dwell-time bounds are inconsistent")
        departures = self.departure_counts(self.baseline_intervals)
        dwell = self.reference_dwell_times(self.baseline_intervals, departures)
        baseline = self._evaluate_raw(self.baseline_intervals, dwell)
        self._scales = {
            "run": max(baseline.run_cost, parameters.epsilon),
            "wait": max(baseline.wait_cost, parameters.epsilon),
            "occ": max(baseline.occupancy_loss, parameters.epsilon),
            "stop": max(baseline.stop_adjustment_cost, 1.0),
        }

    @property
    def decision_dimension(self) -> int:
        return self.period_count + self.period_count * self.station_count

    def bounds(self) -> List[Tuple[float, float]]:
        return list(zip(self.interval_lower, self.interval_upper)) + list(
            zip(self.dwell_lower.ravel(), self.dwell_upper.ravel())
        )

    def vector_to_plan(self, vector: Sequence[float]) -> Tuple[np.ndarray, np.ndarray]:
        values = _as_finite_array(vector, "decision vector").reshape(-1)
        if values.size != self.decision_dimension:
            raise ValueError(f"decision vector must contain {self.decision_dimension} values")
        intervals = np.clip(values[: self.period_count], self.interval_lower, self.interval_upper)
        dwell = values[self.period_count :].reshape(self.period_count, self.station_count)
        return intervals, np.clip(dwell, self.dwell_lower, self.dwell_upper)

    def departure_counts(self, intervals: object) -> np.ndarray:
        """Discrete train counts psi_k = max(1, round(T_k / Delta t_k))."""

        values = _as_finite_array(intervals, "intervals")
        if values.shape != (self.period_count,) or np.any(values <= 0):
            raise ValueError(f"intervals must be positive and have shape ({self.period_count},)")
        return np.maximum(1, np.rint(self.durations / values)).astype(int)

    def reference_dwell_times(self, intervals: np.ndarray, departures: np.ndarray) -> np.ndarray:
        """Passenger-responsive reference dwell times from equation (17)."""

        del intervals  # headways act through the discrete departure counts.
        denominator = departures[:, None] + self.parameters.epsilon
        reference = (
            self.base_dwell[None, :]
            + self.parameters.board_dwell_coeff * self.board_demand / denominator
            + self.parameters.alight_dwell_coeff * self.alight_demand / denominator
        ) * self.stop_pattern
        return np.clip(reference, self.dwell_lower, self.dwell_upper)

    def baseline_vector(self) -> np.ndarray:
        departures = self.departure_counts(self.baseline_intervals)
        dwell = self.reference_dwell_times(self.baseline_intervals, departures)
        return np.concatenate([self.baseline_intervals, dwell.ravel()])

    def heuristic_vector(self) -> np.ndarray:
        demand = np.sum(self.board_demand * self.stop_pattern, axis=1)
        relative = np.maximum(
            demand / max(float(np.mean(demand)), self.parameters.epsilon),
            self.parameters.epsilon,
        )
        intervals = np.clip(
            self.baseline_intervals * np.power(relative, -0.28),
            self.interval_lower,
            self.interval_upper,
        )
        departures = self.departure_counts(intervals)
        dwell = self.reference_dwell_times(intervals, departures)
        return np.concatenate([intervals, dwell.ravel()])

    def passenger_waiting_cost(self, intervals: np.ndarray) -> Tuple[float, float]:
        """Return C_p and the demand-weighted mean wait, equations (11)--(13)."""

        mean_wait = nhpp_mean_waiting_time(intervals, self.parameters.arrival_concentration)
        station_demand = self.board_demand * self.stop_pattern
        passenger_minutes = float(np.sum(station_demand * mean_wait[:, None]))
        total_demand = float(np.sum(station_demand))
        average_wait = passenger_minutes / max(total_demand, self.parameters.epsilon)
        return self.parameters.wait_cost_per_minute * passenger_minutes, average_wait

    def operating_cost(self, departures: np.ndarray) -> float:
        """Return railway operating cost C_q from equation (14)."""

        return float(
            self.parameters.run_cost_per_km
            * self.parameters.line_length_km
            * np.sum(departures)
        )

    def _poisson_target_loss(self, expected: float) -> float:
        """Return 1-F_occ for equations (15)--(16)."""

        target = max(1, int(math.ceil(self.parameters.target_occupancy * self.parameters.train_capacity)))
        if expected <= self.parameters.epsilon:
            return 1.0
        if expected > 30.0:
            z_value = (target - 0.5 - expected) / math.sqrt(expected)
            probability_at_least_target = 1.0 - 0.5 * (1.0 + math.erf(z_value / math.sqrt(2.0)))
            return max(0.0, min(1.0, 1.0 - probability_at_least_target))
        mass = math.exp(-expected)
        cumulative = mass
        for passengers in range(1, target):
            mass *= expected / passengers
            cumulative += mass
        return max(0.0, min(1.0, cumulative))

    def target_occupancy_loss(self, demand: np.ndarray, departures: np.ndarray) -> float:
        return float(sum(
            self._poisson_target_loss(float(demand[index] / departures[index]))
            for index in range(self.period_count)
        ))

    def station_capacities(self, dwell: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        """Return platform and line capacities from equations (19)--(20)."""

        platform = self.platforms[None, :] * self.durations[:, None] / (
            dwell + self.platform_buffer[None, :] + self.parameters.epsilon
        )
        line = self.line_flow[None, :] * self.lines[None, :] * self.durations[:, None]
        return platform, line

    def objective(self, vector: Sequence[float]) -> float:
        intervals, dwell = self.vector_to_plan(vector)
        return self._evaluate_raw(intervals, dwell).objective

    def evaluate_vector(self, vector: Sequence[float]) -> PlanMetrics:
        intervals, dwell = self.vector_to_plan(vector)
        return self._evaluate_raw(intervals, dwell)

    def _evaluate_raw(self, intervals: np.ndarray, dwell_times: np.ndarray) -> PlanMetrics:
        p = self.parameters
        departures = self.departure_counts(intervals)
        reference = self.reference_dwell_times(intervals, departures)
        dwell = np.clip(dwell_times, self.dwell_lower, self.dwell_upper)
        station_demand = self.board_demand * self.stop_pattern
        demand = np.sum(station_demand, axis=1)
        supplied_capacity = p.train_capacity * departures
        served = np.minimum(demand, supplied_capacity)

        wait_cost, average_wait = self.passenger_waiting_cost(intervals)
        run_cost = self.operating_cost(departures)
        occupancy_loss = self.target_occupancy_loss(demand, departures)
        stop_cost = float(np.sum(self.stop_pattern * np.square(dwell - reference)))

        platform_capacity, line_capacity = self.station_capacities(dwell)
        stopped_departures = departures[:, None] * self.stop_pattern
        platform_util = stopped_departures / np.maximum(platform_capacity, p.epsilon)
        line_util = departures[:, None] / np.maximum(line_capacity, p.epsilon)
        combined_util = np.maximum(platform_util, line_util)
        cap_violation = float(np.sum(np.square(np.maximum(0.0, combined_util - 1.0))))

        load = demand / np.maximum(supplied_capacity, p.epsilon)
        load_violation = float(
            np.sum(np.square(np.maximum(0.0, p.load_min - load)))
            + np.sum(np.square(np.maximum(0.0, load - p.load_max)))
        )
        revenue = float(p.average_ticket_fare * np.sum(served))
        revenue_gap = max(0.0, run_cost - revenue)
        demand_gap = np.maximum(0.0, demand - supplied_capacity)
        safe_gap = np.maximum(0.0, p.safe_interval_min - intervals)
        if p.normalize_violations:
            revenue_gap /= max(run_cost, p.epsilon)
            demand_gap = demand_gap / np.maximum(demand, p.epsilon)
            safe_gap = safe_gap / max(p.safe_interval_min, p.epsilon)
        revenue_violation = float(revenue_gap**2)
        demand_violation = float(np.sum(np.square(demand_gap)))
        safe_violation = float(np.sum(np.square(safe_gap)))

        total_demand = float(np.sum(demand))
        served_demand = float(np.sum(served))
        weighted_interval = float(np.sum(demand * intervals) / max(total_demand, p.epsilon))
        scales = getattr(self, "_scales", {"run": 1.0, "wait": 1.0, "occ": 1.0, "stop": 1.0})
        if not p.normalize_objectives:
            scales = {"run": 1.0, "wait": 1.0, "occ": 1.0, "stop": 1.0}
        objective = (
            p.omega_run_cost * run_cost / scales["run"]
            + p.omega_wait_cost * wait_cost / scales["wait"]
            + p.omega_occ_loss * occupancy_loss / scales["occ"]
            + p.omega_stop_cost * stop_cost / scales["stop"]
            + p.penalty_cap * cap_violation
            + p.penalty_load * load_violation
            + p.penalty_revenue * revenue_violation
            + p.penalty_demand * demand_violation
            + p.penalty_safe * safe_violation
        )
        return PlanMetrics(
            intervals.copy(), departures.copy(), dwell.copy(), reference.copy(), run_cost, wait_cost,
            occupancy_loss, stop_cost, revenue, total_demand, served_demand,
            served_demand / max(total_demand, p.epsilon), weighted_interval, average_wait,
            served_demand / max(float(np.sum(supplied_capacity)), p.epsilon),
            float(np.max(combined_util)), cap_violation, load_violation, revenue_violation,
            demand_violation, safe_violation, float(objective), platform_capacity, line_capacity,
            platform_util, line_util, combined_util,
        )


# Backward-compatible name used by existing experiment utilities.
HSRModel = JointHeadwayDwellModel


class ImprovedModel:
    def __init__(self, base_model: JointHeadwayDwellModel):
        self.base = base_model

    def bounds(self) -> List[Tuple[float, float]]:
        return self.base.bounds()

    def baseline_vector(self) -> np.ndarray:
        return self.base.baseline_vector()

    def heuristic_vector(self) -> np.ndarray:
        return self.base.heuristic_vector()

    def seed_solutions(self) -> List[List[float]]:
        baseline, heuristic = self.baseline_vector(), self.heuristic_vector()
        midpoint = 0.35 * baseline + 0.65 * heuristic
        aggressive = heuristic.copy()
        aggressive[: self.base.period_count] = np.clip(
            aggressive[: self.base.period_count] * 0.97,
            self.base.interval_lower,
            self.base.interval_upper,
        )
        return [baseline.tolist(), heuristic.tolist(), midpoint.tolist(), aggressive.tolist()]

    def repair_vector(self, vector: Sequence[float]) -> np.ndarray:
        intervals, dwell = self.base.vector_to_plan(vector)
        intervals, dwell = intervals.copy(), dwell.copy()
        for _ in range(8):
            departures = self.base.departure_counts(intervals)
            dwell = np.clip(
                np.minimum(dwell, self.base.reference_dwell_times(intervals, departures)),
                self.base.dwell_lower,
                self.base.dwell_upper,
            )
            platform, line = self.base.station_capacities(dwell)
            platform_allowable = np.where(
                self.base.stop_pattern > 0,
                np.floor(platform + 1.0e-9),
                np.inf,
            )
            allowable = np.clip(
                np.floor(np.min(np.minimum(platform_allowable, line), axis=1) + 1.0e-9).astype(int),
                1,
                None,
            )
            new_departures = np.minimum(departures, allowable)
            if np.array_equal(new_departures, departures):
                break
            intervals = np.clip(
                self.base.durations / new_departures,
                self.base.interval_lower,
                self.base.interval_upper,
            )
        return np.concatenate([intervals, dwell.ravel()])

    def objective(self, vector: Sequence[float]) -> float:
        return float(self.base.objective(vector))

    def evaluate_vector(self, vector: Sequence[float]) -> PlanMetrics:
        return self.base.evaluate_vector(self.repair_vector(vector))


def build_dataset_model(
    code: str = "NST-HSR",
    parameter_overrides: Mapping[str, object] | None = None,
) -> ImprovedModel:
    """Build the canonical model from a standardized dataset bundle."""

    if code not in DATASET_SPECS:
        raise KeyError(f"Unknown dataset {code!r}; available values: {', '.join(DATASET_SPECS)}")
    timetable, factors, capacity = find_bundle_files(DATASET_SPECS[code])
    rows = load_timetable(timetable)
    periods = build_periods()
    stations = load_station_capacities(capacity)
    board, alight = load_factor_demands(factors, periods, stations)
    capacities = [float(row["train_capacity"]) for row in rows if float(row["train_capacity"]) > 0]
    fares = [float(row["second_class_fare"]) for row in rows if float(row["second_class_fare"]) > 0]
    parameters = ModelParameters(
        train_capacity=int(round(median(capacities))) if capacities else 600,
        line_length_km=max(station.km for station in stations),
        average_ticket_fare=float(mean(fares)) if fares else 12.0,
    )
    if parameter_overrides:
        parameters = replace(parameters, **dict(parameter_overrides))
    return ImprovedModel(JointHeadwayDwellModel(
        periods,
        stations,
        compute_period_statistics(rows, periods),
        board,
        alight,
        parameters,
    ))


def create_scenario_model(
    model: ImprovedModel,
    scenario: str | ScenarioDefinition,
) -> ImprovedModel:
    definition = SCENARIOS[scenario] if isinstance(scenario, str) else scenario
    base = model.base
    station_scale = np.array([
        definition.station_focus.get(index, 1.0)
        for index, _ in enumerate(base.stations)
    ])
    period_scale = np.asarray(definition.period_scale, dtype=float)[:, None]
    if period_scale.shape != (base.period_count, 1):
        raise ValueError(f"scenario period_scale must contain {base.period_count} values")
    board = base.board_demand * period_scale * station_scale[None, :]
    alight = base.alight_demand * period_scale * station_scale[None, :]
    target = float(np.sum(base.board_demand) * definition.overall_scale)
    board *= target / max(float(np.sum(board)), base.parameters.epsilon)
    alight *= target / max(float(np.sum(alight)), base.parameters.epsilon)
    return ImprovedModel(JointHeadwayDwellModel(
        base.periods,
        base.stations,
        base.period_stats,
        board,
        alight,
        base.parameters,
        stop_pattern=base.stop_pattern,
    ))


def hard_violation(metrics: PlanMetrics) -> float:
    return max(0.0, metrics.max_capacity_utilization - 1.0) + metrics.safe_violation


def soft_violation(metrics: PlanMetrics) -> float:
    return metrics.demand_violation + 0.5 * metrics.load_violation + metrics.revenue_violation


def benchmark_score(metrics: PlanMetrics) -> float:
    return metrics.objective + 1.0e5 * hard_violation(metrics) + 1.0e3 * soft_violation(metrics)


def benchmark_rank_key(metrics: PlanMetrics) -> tuple:
    return (
        hard_violation(metrics) > 1.0e-9,
        hard_violation(metrics),
        soft_violation(metrics),
        metrics.objective,
        metrics.average_waiting_time,
    )


__all__ = [
    "DATASET_SPECS",
    "SCENARIOS",
    "DatasetSpec",
    "GeneralizedCostCoefficients",
    "HSRModel",
    "ImprovedModel",
    "JointHeadwayDwellModel",
    "ModelParameters",
    "PassengerDemandModel",
    "PeriodDefinition",
    "PeriodStats",
    "PlanMetrics",
    "ScenarioDefinition",
    "StationCapacity",
    "allocate_rail_demand",
    "benchmark_rank_key",
    "benchmark_score",
    "build_dataset_model",
    "build_periods",
    "compute_period_statistics",
    "create_scenario_model",
    "cumulative_arrival_intensity",
    "find_bundle_files",
    "generalized_cost",
    "gm11_forecast",
    "hard_violation",
    "induced_demand",
    "integrate_interval_waiting_time",
    "load_factor_demands",
    "load_station_capacities",
    "load_timetable",
    "logit_mode_probabilities",
    "map_od_demand_to_period_station",
    "nhpp_mean_waiting_time",
    "normalize_station_name",
    "power_arrival_intensity",
    "soft_violation",
    "to_minutes",
]
